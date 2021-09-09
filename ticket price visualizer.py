import openpyxl
import matplotlib.pyplot as plt
from openpyxl.cell import cell

#create workbook
wb = openpyxl.load_workbook("C:/Users/aweso/Desktop/python projects/nasa/flighticket project/Average Domestic Airline Itinerary Fares q4 NASA.xlsx")
#define data record timeline
firstYearRecorded = 1993
lastYearRecorded = 2019

#returns airport avg prices for all years in list inflated and uninflated
def getAllAnnualPricesAtAirport(airportcode,isInflated):
    #creating list of prices for all years at airport (uninflated and inflated)
    prices = list()
    iteratingYear = firstYearRecorded
    while iteratingYear<=lastYearRecorded:
        prices.append(getPriceAtAirport(iteratingYear,airportcode,isInflated))
        iteratingYear+=1
    return prices

#returns national avg prices for all years in list inflated and uninflated
def getAllAnnualNationalAvgPrices(isInflated):
    #creating list of national prices for all years (uninflated and inflated)
    prices = list()
    iteratingYear = firstYearRecorded
    while iteratingYear<=lastYearRecorded:
        prices.append(getNationalPriceFromAnalysis(iteratingYear,isInflated))
        iteratingYear+=1
    return prices

#returns airport avg fare uninflated or inflated
def getPriceAtAirport(year,airportcode,isInflated):
    sheet = wb[str(year)]
    rowIterator = 1
    lastRowInSheet = sheet.max_row-2
    if isInflated:
        col = 6
    else:
        col = 5
    while rowIterator <= lastRowInSheet:
        cellData = sheet.cell(rowIterator,2).value
        if airportcode == cellData:
            break
        rowIterator+=1
    price = sheet.cell(rowIterator,col).value
    return price

#returns national avg fare uninflated or inflated
def getNationalPrice(year,isInflated):
    sheet = wb[str(year)]
    row = sheet.max_row-1
    if isInflated:
        col = 6
    else:
        col = 5
    price = sheet.cell(row,col).value
    return price

def getNationalPriceFromAnalysis(year,isInflated):
    sheet = wb['analysis']
    row = (year - firstYearRecorded) + 2
    if isInflated:
        col = 3
    else:
        col = 2
    price = sheet.cell(row,col).value
    return price

#returns percent change between two values (formatting depending on growth or decline)
def percentChange(finalValue, initialValue):
    change = (finalValue-initialValue)/initialValue
    change *= 100
    changeString = str(change)
    if change > 0:
        changeString = '+'+changeString
    return changeString+'%'

def computePercentChanges():
    priorYear = input('Enter first year: ')
    laterYear = input('Enter later year: ')
    airportcode = input('Enter airport code: ')
    airportcode = airportcode.upper()

    if airportcode != 'national':
        #calculate percent difference in avg fare at specific airport (no inflation)
        percentChange_AtAirport_notInflated = percentChange(getPriceAtAirport(int(laterYear),airportcode,False),getPriceAtAirport(int(priorYear),airportcode,False))
        print("Percent Change in avg ticket fare (uninflated) at airport "+airportcode+": "+(percentChange_AtAirport_notInflated))

        #calculate percent difference in avg fare at specific airport (with inflation)
        percentChange_AtAirport_Inflated = percentChange(getPriceAtAirport(int(laterYear),airportcode,True),getPriceAtAirport(int(priorYear),airportcode,True))
        print("Percent Change in avg ticket fare (inflated) at airport "+airportcode+": "+(percentChange_AtAirport_Inflated))

        #calculate percent difference in avg fare nationally (no inflation)
        percentChange_Nationally_notInflated = percentChange(getNationalPrice(int(laterYear),False),getNationalPrice(int(priorYear),False))
        print("Percent Change in national avg ticket fare (uninflated): "+(percentChange_Nationally_notInflated))

        #calculate percent difference in avg fare nationally (inflation)
        percentChange_Nationally_Inflated = percentChange(getNationalPrice(int(laterYear),True),getNationalPrice(int(priorYear),True))
        print("Percent Change in national avg ticket fare (inflated): "+(percentChange_Nationally_Inflated))
    else:
        #calculate percent difference in avg fare nationally (no inflation)
        percentChange_Nationally_notInflated = percentChange(getNationalPrice(int(laterYear),False),getNationalPrice(int(priorYear),False))
        print("Percent Change in national avg ticket fare (uninflated): "+(percentChange_Nationally_notInflated))

        #calculate percent difference in avg fare nationally (inflation)
        percentChange_Nationally_Inflated = percentChange(getNationalPrice(int(laterYear),True),getNationalPrice(int(priorYear),True))
        print("Percent Change in national avg ticket fare (inflated): "+(percentChange_Nationally_Inflated))

#graphing prices over time (inflated is orange, uninflated is blue)
def showPricesOverTime():
    airportcode = input('Enter airport code: ')
    airportcode = airportcode.upper()
    # x axis values
    x = list()
    for year in range(27):
        x.append(year+firstYearRecorded)

    #specific airport prices
    if airportcode != 'NATIONAL':
        # corresponding y axis values (uninflated)
        y_uninflated = getAllAnnualPricesAtAirport(airportcode,False)
        # corresponding y axis values (inflated)
        y_inflated = getAllAnnualPricesAtAirport(airportcode,True)
        
        # plotting the points 
        plt.plot(x, y_uninflated)
        plt.plot(x, y_inflated)
        
        # naming the x axis
        plt.xlabel('Year')
        # naming the y axis
        plt.ylabel('Price in Dollars')
    
        # giving a title to my graph
        plt.title('Avg Ticket Fare at Airport '+airportcode+' Over Time\n Orange Line: Prices Based on Inflation\n Blue Line: Prices not Based on Inflation')
        
        # function to show the plot
        plt.show()
    #national airport prices
    else:
        # corresponding y axis values (uninflated)
        y_uninflated = getAllAnnualNationalAvgPrices(False)
        # corresponding y axis values (inflated)
        y_inflated = getAllAnnualNationalAvgPrices(True)
        
        # plotting the points 
        plt.plot(x, y_uninflated)
        plt.plot(x, y_inflated)
        
        # naming the x axis
        plt.xlabel('Year')
        # naming the y axis
        plt.ylabel('Price in Dollars')
    
        # giving a title to my graph
        plt.title('National Avg Ticket Fare Over Time\n Orange Line: Prices Based on Inflation\n Blue Line: Prices not Based on Inflation')
        
        # function to show the plot
        plt.show()


#   --main--
#ask user to see graphs or specific changes
while True:
    print('1. See ticket prices at specific years')
    print('2. See graphs')
    print('3. Quit')
    while True:
        option = input('Enter option: ')
        choice = int(option)
        if choice == 1 or choice ==2 or choice == 3:
            break

    #give appropriate output based on user choice
    if choice == 1:
        computePercentChanges()
    elif choice == 2:
        showPricesOverTime()
    else:
        break
